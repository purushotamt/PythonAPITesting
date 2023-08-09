import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_student():
    #API
    API_URL = 'https://thetestingworldapi.com/api/studentsDetails'
    f=open('C:/Users/Purushotam/Desktop/AddNewstudent.json')
    #excel code
    wk = openpyxl.load_workbook('C:/Users/Purushotam/Desktop/TestDAT.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row
    json_request = json.loads(f.read())
    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i,column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value
        response = requests.post(API_URL, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201











